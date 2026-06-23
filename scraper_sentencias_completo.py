"""
Scraper completo - Sentencias Control de Constitucionalidad
Corte Constitucional del Ecuador
Extrae contenido de cada sentencia y genera Excel.
Requiere: pip install requests beautifulsoup4 lxml openpyxl
"""

import requests
from bs4 import BeautifulSoup
import time
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-EC,es;q=0.9,en;q=0.8",
    "Referer": "https://www.google.com/",
}

session = requests.Session()
session.headers.update(HEADERS)

TERMINOS_BUSQUEDA = [
    "sentencia control constitucionalidad",
    "inconstitucionalidad abstracto Ecuador",
    "consulta norma control concreto Ecuador",
    "accion inconstitucionalidad corte constitucional Ecuador",
    "referendo control constitucional Ecuador",
    "control abstracto constitucionalidad Ecuador sentencia",
]


def buscar_google(termino: str) -> list[dict]:
    resultados = []
    url = f"https://www.google.com/search?q=site:corteconstitucional.gob.ec+{termino.replace(' ', '+')}&num=10&hl=es"
    try:
        r = session.get(url, timeout=15)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, "lxml")
            for div in soup.find_all("div", class_=["tF2Cxc", "g"]):
                a = div.find("a", href=True)
                h3 = div.find("h3")
                desc = div.find(["span", "div"], class_=["VwiC3b", "aCOpRe", "lEBKkf"])
                if a and "corteconstitucional.gob.ec" in a["href"]:
                    resultados.append({
                        "titulo": h3.get_text(strip=True) if h3 else "",
                        "url": a["href"],
                        "descripcion_google": desc.get_text(strip=True)[:300] if desc else "",
                    })
    except Exception as e:
        print(f"  [Google error] {e}")
    return resultados


def buscar_bing(termino: str) -> list[dict]:
    resultados = []
    url = f"https://www.bing.com/search?q=site:corteconstitucional.gob.ec+{termino.replace(' ', '+')}&count=10"
    try:
        r = session.get(url, timeout=15)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, "lxml")
            for li in soup.find_all("li", class_="b_algo"):
                a = li.find("a", href=True)
                h2 = li.find("h2")
                p = li.find("p")
                if a and "corteconstitucional.gob.ec" in a.get("href", ""):
                    resultados.append({
                        "titulo": h2.get_text(strip=True) if h2 else "",
                        "url": a["href"],
                        "descripcion_google": p.get_text(strip=True)[:300] if p else "",
                    })
    except Exception as e:
        print(f"  [Bing error] {e}")
    return resultados


def extraer_contenido_sentencia(url: str) -> dict:
    """Entra a la página de la sentencia y extrae su contenido."""
    datos = {
        "numero_sentencia": "",
        "tipo_accion": "",
        "fecha": "",
        "juez_ponente": "",
        "accionante": "",
        "accionado": "",
        "norma_impugnada": "",
        "objeto": "",
        "procedimiento": "",
        "decision": "",
        "resolucion": "",
        "texto_completo": "",
    }
    try:
        r = session.get(url, timeout=20)
        if r.status_code != 200:
            return datos
        soup = BeautifulSoup(r.text, "lxml")

        # Texto completo de la página
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        texto = soup.get_text(separator="\n", strip=True)
        datos["texto_completo"] = texto[:3000]

        # Número de sentencia (patrones comunes: 88-22-IN/25, 001-13-SCN-CC)
        patron_num = re.search(
            r'\d{1,4}-\d{2}-(?:IN|CN|RC|SCN|SIN|EP|EE|AA|AP|AN|CC|IS|RE|RA|UP|MP)/?\d{0,2}[-\w]*',
            texto, re.IGNORECASE
        )
        if patron_num:
            datos["numero_sentencia"] = patron_num.group(0)

        # Tipo de acción
        tipos = {
            "IN": "Acción de Inconstitucionalidad (Control Abstracto)",
            "CN": "Consulta de Norma (Control Concreto)",
            "RC": "Control Constitucionalidad de Referendo",
            "SCN": "Sentencia de Consulta de Norma",
            "SIN": "Sentencia de Inconstitucionalidad",
            "EE": "Control Constitucionalidad de Decreto",
        }
        for cod, desc in tipos.items():
            if f"-{cod}" in (datos["numero_sentencia"] or texto[:500]).upper():
                datos["tipo_accion"] = desc
                break

        # Fecha
        patron_fecha = re.search(
            r'(\d{1,2}\s+de\s+\w+\s+de\s+20\d{2}|\d{2}/\d{2}/20\d{2})',
            texto, re.IGNORECASE
        )
        if patron_fecha:
            datos["fecha"] = patron_fecha.group(0)

        # Juez ponente
        patron_juez = re.search(
            r'(?:juez[a]?\s+(?:constitucional\s+)?ponente|sustanciador[a]?)[:\s]+([A-ZÁÉÍÓÚÑ][a-záéíóúñ\s]+)',
            texto, re.IGNORECASE
        )
        if patron_juez:
            datos["juez_ponente"] = patron_juez.group(1).strip()[:100]

        # Norma impugnada
        patron_norma = re.search(
            r'(?:norma\s+(?:impugnada|demandada|cuestionada)|objeto\s+de\s+control)[:\s]+(.{20,200})',
            texto, re.IGNORECASE
        )
        if patron_norma:
            datos["norma_impugnada"] = patron_norma.group(1).strip()[:300]

        # Objeto / materia
        patron_objeto = re.search(
            r'(?:objeto|materia|pretensión)[:\s]+(.{20,300})',
            texto, re.IGNORECASE
        )
        if patron_objeto:
            datos["objeto"] = patron_objeto.group(1).strip()[:300]

        # Procedimiento (buscar sección)
        patron_proc = re.search(
            r'(?:procedimiento|trámite|sustanciación|antecedentes)[:\s\n]+(.{50,500})',
            texto, re.IGNORECASE
        )
        if patron_proc:
            datos["procedimiento"] = patron_proc.group(1).strip()[:500]

        # Decisión / resolución
        patron_res = re.search(
            r'(?:resuelve|por\s+lo\s+expuesto|en\s+mérito\s+de\s+lo\s+expuesto)[:\s\n]+(.{50,500})',
            texto, re.IGNORECASE
        )
        if patron_res:
            datos["resolucion"] = patron_res.group(1).strip()[:500]

        # Título de la página como respaldo
        title = soup.find("title")
        if title and not datos["numero_sentencia"]:
            datos["numero_sentencia"] = title.get_text(strip=True)[:100]

    except Exception as e:
        datos["texto_completo"] = f"Error al acceder: {e}"

    return datos


def generar_excel(sentencias: list[dict], archivo: str = "sentencias_control_constitucionalidad.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sentencias CC"

    # Estilos
    color_header = "1F4E79"
    color_fila_par = "D6E4F0"
    color_fila_impar = "FFFFFF"

    font_header = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_normal = Font(name="Calibri", size=10)
    alin_centro = Alignment(horizontal="center", vertical="top", wrap_text=True)
    alin_izq = Alignment(horizontal="left", vertical="top", wrap_text=True)
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    columnas = [
        ("N°",                      8),
        ("Número Sentencia",        22),
        ("Tipo de Acción",          30),
        ("Fecha",                   18),
        ("Juez Ponente",            25),
        ("Norma Impugnada",         35),
        ("Objeto / Materia",        40),
        ("Procedimiento",           50),
        ("Resolución",              50),
        ("URL",                     50),
        ("Descripción (buscador)",  40),
    ]

    # Encabezados
    for col_idx, (nombre, ancho) in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=nombre)
        cell.font = font_header
        cell.fill = PatternFill("solid", fgColor=color_header)
        cell.alignment = alin_centro
        cell.border = borde
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Datos
    for i, s in enumerate(sentencias, 1):
        fila = i + 1
        color = color_fila_par if i % 2 == 0 else color_fila_impar
        fill = PatternFill("solid", fgColor=color)

        valores = [
            i,
            s.get("numero_sentencia", ""),
            s.get("tipo_accion", ""),
            s.get("fecha", ""),
            s.get("juez_ponente", ""),
            s.get("norma_impugnada", ""),
            s.get("objeto", ""),
            s.get("procedimiento", ""),
            s.get("resolucion", ""),
            s.get("url", ""),
            s.get("descripcion_google", ""),
        ]

        for col_idx, valor in enumerate(valores, 1):
            cell = ws.cell(row=fila, column=col_idx, value=str(valor) if valor else "")
            cell.font = font_normal
            cell.fill = fill
            cell.border = borde
            cell.alignment = alin_centro if col_idx == 1 else alin_izq

        ws.row_dimensions[fila].height = 60

    # Segunda hoja: texto completo
    ws2 = wb.create_sheet("Texto Completo")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 120

    headers2 = ["Número Sentencia", "URL", "Texto Extraído (primeros 3000 chars)"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = PatternFill("solid", fgColor=color_header)
        cell.alignment = alin_centro
        cell.border = borde

    for i, s in enumerate(sentencias, 2):
        ws2.cell(row=i, column=1, value=s.get("numero_sentencia", "")).alignment = alin_izq
        ws2.cell(row=i, column=2, value=s.get("url", "")).alignment = alin_izq
        ws2.cell(row=i, column=3, value=s.get("texto_completo", "")).alignment = alin_izq
        ws2.row_dimensions[i].height = 100

    wb.save(archivo)
    print(f"\n[OK] Excel generado: {archivo}")
    print(f"     {len(sentencias)} sentencias en 2 hojas (resumen + texto completo)")


def main():
    print("=" * 65)
    print("  Scraper Completo - Control de Constitucionalidad")
    print("  Corte Constitucional del Ecuador")
    print("=" * 65 + "\n")

    # 1. Recopilar URLs
    print("[1] Buscando sentencias en Google y Bing...\n")
    todos = []
    vistos = set()

    for termino in TERMINOS_BUSQUEDA:
        print(f"  Término: {termino}")
        for item in buscar_google(termino) + buscar_bing(termino):
            url = item.get("url", "")
            if url and url not in vistos and "corteconstitucional.gob.ec" in url:
                vistos.add(url)
                todos.append(item)
        time.sleep(2)

    print(f"\n  => {len(todos)} URLs únicas encontradas\n")

    # 2. Extraer contenido de cada sentencia
    print("[2] Extrayendo contenido de cada sentencia...\n")
    sentencias = []
    for i, item in enumerate(todos, 1):
        url = item["url"]
        print(f"  [{i}/{len(todos)}] {url[:80]}")
        contenido = extraer_contenido_sentencia(url)
        contenido["url"] = url
        contenido["descripcion_google"] = item.get("descripcion_google", "")
        if not contenido["numero_sentencia"]:
            contenido["numero_sentencia"] = item.get("titulo", "")[:100]
        sentencias.append(contenido)
        time.sleep(1.5)

    # 3. Generar Excel
    print("\n[3] Generando archivo Excel...\n")
    generar_excel(sentencias)

    print("\nAbre el archivo:")
    print("  sentencias_control_constitucionalidad.xlsx")


if __name__ == "__main__":
    main()
