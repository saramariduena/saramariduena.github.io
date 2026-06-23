"""
Scraper completo - Sentencias Control de Constitucionalidad
Corte Constitucional del Ecuador - www.corteconstitucional.gob.ec
Busca directamente en el sitio oficial y genera Excel.
Requiere: pip install requests beautifulsoup4 lxml openpyxl
"""

import requests
from bs4 import BeautifulSoup
import time, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin

BASE = "https://www.corteconstitucional.gob.ec"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-EC,es;q=0.9,en;q=0.8",
}

TERMINOS = [
    "control de constitucionalidad",
    "accion de inconstitucionalidad",
    "consulta de norma",
    "control abstracto",
    "control concreto",
]

session = requests.Session()
session.headers.update(HEADERS)


def buscar_en_sitio(termino: str, max_paginas: int = 5) -> list[dict]:
    """Usa el buscador interno de www.corteconstitucional.gob.ec"""
    resultados = []
    print(f"\n  Buscando: '{termino}'")

    for pagina in range(1, max_paginas + 1):
        if pagina == 1:
            url = f"{BASE}/?s={termino.replace(' ', '+')}"
        else:
            url = f"{BASE}/page/{pagina}/?s={termino.replace(' ', '+')}"

        try:
            r = session.get(url, timeout=20)
            if r.status_code == 404:
                break
            if r.status_code != 200:
                print(f"    [Página {pagina}] Status {r.status_code}")
                break

            soup = BeautifulSoup(r.text, "lxml")

            # Buscar artículos/entradas del blog WordPress
            articulos = soup.find_all(["article", "div"], class_=re.compile(r"post|entry|result|item", re.I))

            if not articulos:
                # Buscar enlaces directamente
                articulos = soup.find_all("h2")

            encontrados = 0
            for art in articulos:
                a = art.find("a", href=True)
                if not a:
                    continue
                href = a["href"]
                if BASE not in href and not href.startswith("/"):
                    continue
                titulo = a.get_text(strip=True)
                if not titulo:
                    h = art.find(["h1","h2","h3","h4"])
                    titulo = h.get_text(strip=True) if h else href

                # Descripción/extracto
                desc_tag = art.find(["p", "div"], class_=re.compile(r"excerpt|summary|description|entry-summary", re.I))
                descripcion = desc_tag.get_text(strip=True)[:300] if desc_tag else ""

                url_sentencia = href if href.startswith("http") else urljoin(BASE, href)

                if url_sentencia not in [x["url"] for x in resultados]:
                    resultados.append({
                        "titulo": titulo[:150],
                        "url": url_sentencia,
                        "descripcion": descripcion,
                        "termino_busqueda": termino,
                    })
                    encontrados += 1

            print(f"    Página {pagina}: {encontrados} resultados")

            if encontrados == 0:
                break

            time.sleep(1.5)

        except Exception as e:
            print(f"    [Error página {pagina}] {e}")
            break

    return resultados


def extraer_sentencia(url: str) -> dict:
    """Extrae el contenido detallado de una sentencia."""
    datos = {
        "numero_sentencia": "",
        "tipo_accion": "",
        "fecha": "",
        "juez_ponente": "",
        "norma_impugnada": "",
        "objeto": "",
        "antecedentes": "",
        "procedimiento": "",
        "resolucion": "",
        "texto_pagina": "",
    }

    try:
        r = session.get(url, timeout=20)
        if r.status_code != 200:
            datos["texto_pagina"] = f"Error HTTP {r.status_code}"
            return datos

        soup = BeautifulSoup(r.text, "lxml")

        # Remover navegación, scripts, etc.
        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()

        # Contenido principal
        contenido = (
            soup.find("div", class_=re.compile(r"entry-content|post-content|content|article", re.I))
            or soup.find("main")
            or soup.find("article")
            or soup.body
        )

        texto = contenido.get_text(separator="\n", strip=True) if contenido else ""
        datos["texto_pagina"] = texto[:4000]

        # Número de sentencia
        patron = re.search(
            r'\d{1,4}-\d{2}-(?:IN|CN|RC|SCN|SIN|EE|EP|IS|RE|AA|AN|AP|CC|RA|UP|MP)[-/]?\d{0,2}',
            texto, re.IGNORECASE
        )
        if patron:
            datos["numero_sentencia"] = patron.group(0)
        else:
            # Tomar del título de la página
            title = soup.find("h1") or soup.find("title")
            if title:
                datos["numero_sentencia"] = title.get_text(strip=True)[:120]

        # Tipo de acción
        tipos = {
            "-IN": "Inconstitucionalidad (Control Abstracto)",
            "-CN": "Consulta de Norma (Control Concreto)",
            "-RC": "Control Constitucionalidad de Referendo",
            "-SCN": "Sentencia Consulta de Norma",
            "-SIN": "Sentencia de Inconstitucionalidad",
            "-EE": "Control de Decreto Ejecutivo",
        }
        num = datos["numero_sentencia"].upper()
        for cod, desc in tipos.items():
            if cod in num:
                datos["tipo_accion"] = desc
                break

        # Fecha
        m = re.search(r'\d{1,2}\s+de\s+\w+\s+de\s+20\d{2}', texto, re.IGNORECASE)
        if m:
            datos["fecha"] = m.group(0)

        # Juez ponente
        m = re.search(
            r'(?:juez[a]?\s+(?:constitucional\s+)?ponente|sustanciador[a]?)[:\s]+([A-ZÁÉÍÓÚÑ][^\n]{5,80})',
            texto, re.IGNORECASE
        )
        if m:
            datos["juez_ponente"] = m.group(1).strip()[:100]

        # Norma impugnada
        m = re.search(
            r'(?:norma\s+(?:impugnada|demandada|cuestionada))[:\s]+([^\n]{20,300})',
            texto, re.IGNORECASE
        )
        if m:
            datos["norma_impugnada"] = m.group(1).strip()[:300]

        # Objeto
        m = re.search(
            r'(?:objeto|materia|pretensión)[:\s]+([^\n]{20,300})',
            texto, re.IGNORECASE
        )
        if m:
            datos["objeto"] = m.group(1).strip()[:300]

        # Antecedentes
        m = re.search(
            r'(?:antecedentes|i\.\s*antecedentes)[:\s\n]+(.{100,800}?)(?:\n\n|\Z)',
            texto, re.IGNORECASE | re.DOTALL
        )
        if m:
            datos["antecedentes"] = m.group(1).strip()[:600]

        # Procedimiento
        m = re.search(
            r'(?:procedimiento|trámite|sustanciación)[:\s\n]+(.{50,600}?)(?:\n\n|\Z)',
            texto, re.IGNORECASE | re.DOTALL
        )
        if m:
            datos["procedimiento"] = m.group(1).strip()[:500]

        # Resolución
        m = re.search(
            r'(?:resuelve|por\s+lo\s+expuesto|en\s+mérito)[:\s\n]+(.{50,600}?)(?:\n\n|\Z)',
            texto, re.IGNORECASE | re.DOTALL
        )
        if m:
            datos["resolucion"] = m.group(1).strip()[:500]

    except Exception as e:
        datos["texto_pagina"] = f"Error: {e}"

    return datos


def generar_excel(sentencias: list[dict], archivo="sentencias_control_constitucionalidad.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Sentencias"

    azul = "1F4E79"
    fnt_h = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fnt   = Font(name="Calibri", size=10)
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    al_c = Alignment(horizontal="center", vertical="top", wrap_text=True)
    al_l = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    columnas = [
        ("N°",                    6),
        ("Número Sentencia",     25),
        ("Tipo de Acción",       35),
        ("Fecha",                20),
        ("Juez Ponente",         28),
        ("Norma Impugnada",      38),
        ("Objeto / Materia",     40),
        ("Antecedentes",         50),
        ("Procedimiento",        50),
        ("Resolución",           50),
        ("URL",                  55),
        ("Término de Búsqueda", 30),
    ]

    for ci, (nombre, ancho) in enumerate(columnas, 1):
        c = ws.cell(row=1, column=ci, value=nombre)
        c.font = fnt_h
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = al_c
        c.border = borde
        ws.column_dimensions[get_column_letter(ci)].width = ancho

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, s in enumerate(sentencias, 1):
        fila = i + 1
        fill = PatternFill("solid", fgColor="D6E4F0" if i % 2 == 0 else "FFFFFF")
        vals = [
            i,
            s.get("numero_sentencia", ""),
            s.get("tipo_accion", ""),
            s.get("fecha", ""),
            s.get("juez_ponente", ""),
            s.get("norma_impugnada", ""),
            s.get("objeto", ""),
            s.get("antecedentes", ""),
            s.get("procedimiento", ""),
            s.get("resolucion", ""),
            s.get("url", ""),
            s.get("termino_busqueda", ""),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=fila, column=ci, value=str(v) if v else "")
            c.font = fnt
            c.fill = fill
            c.border = borde
            c.alignment = al_c if ci == 1 else al_l
        ws.row_dimensions[fila].height = 70

    # Hoja 2: texto completo
    ws2 = wb.create_sheet("Texto Completo")
    for ci, (h, w) in enumerate([("Número", 25), ("URL", 50), ("Texto extraído", 120)], 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = fnt_h
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = al_c
        c.border = borde
        ws2.column_dimensions[get_column_letter(ci)].width = w

    for i, s in enumerate(sentencias, 2):
        ws2.cell(row=i, column=1, value=s.get("numero_sentencia", "")).alignment = al_l
        ws2.cell(row=i, column=2, value=s.get("url", "")).alignment = al_l
        ws2.cell(row=i, column=3, value=s.get("texto_pagina", "")).alignment = al_l
        ws2.row_dimensions[i].height = 120

    wb.save(archivo)
    print(f"\n[OK] Excel generado: {archivo}")
    print(f"     {len(sentencias)} sentencias — 2 hojas: Resumen + Texto Completo")


def main():
    print("=" * 65)
    print("  Scraper - www.corteconstitucional.gob.ec")
    print("  Control de Constitucionalidad")
    print("=" * 65)

    # Verificar acceso
    print("\n[0] Verificando acceso al sitio...")
    try:
        r = session.get(BASE, timeout=15)
        print(f"    [OK] Sitio accesible (HTTP {r.status_code})")
    except Exception as e:
        print(f"    [ERROR] No se puede acceder: {e}")
        print("    Verifica tu conexión a internet.")
        return

    # Buscar en el sitio
    print("\n[1] Buscando sentencias en el sitio oficial...")
    urls_encontradas = {}
    for termino in TERMINOS:
        for item in buscar_en_sitio(termino, max_paginas=5):
            url = item["url"]
            if url not in urls_encontradas:
                urls_encontradas[url] = item

    total = len(urls_encontradas)
    print(f"\n  => {total} páginas únicas encontradas")

    if total == 0:
        print("  No se encontraron resultados. Intenta abrir manualmente:")
        print(f"  {BASE}/?s=control+de+constitucionalidad")
        return

    # Extraer contenido de cada sentencia
    print("\n[2] Extrayendo contenido de cada página...")
    sentencias = []
    for i, (url, meta) in enumerate(urls_encontradas.items(), 1):
        print(f"  [{i}/{total}] {url[:80]}")
        contenido = extraer_sentencia(url)
        contenido["url"] = url
        contenido["termino_busqueda"] = meta.get("termino_busqueda", "")
        if not contenido["numero_sentencia"]:
            contenido["numero_sentencia"] = meta.get("titulo", "")[:120]
        sentencias.append(contenido)
        time.sleep(1)

    # Generar Excel
    print("\n[3] Generando Excel...")
    generar_excel(sentencias)
    print("\nAbre el archivo: sentencias_control_constitucionalidad.xlsx")


if __name__ == "__main__":
    main()
